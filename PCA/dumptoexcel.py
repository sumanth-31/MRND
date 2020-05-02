from bs4 import BeautifulSoup
import click
from openpyxl import Workbook
from os.path import isfile
from openpyxl.styles import Font
@click.command()
@click.argument('src')
@click.argument('dest')
def dumptoexcel(src,dest):
    '''Command Line tool to copy table data from html file to excel sheet'''
    if(isfile(dest)):
        if(not click.confirm('File already exists. Do you want to overwrite?')):
            return
    with open(src) as fp:
        soup=BeautifulSoup(fp,features='lxml')
        wb=Workbook()
        ws=wb.active
        click.echo('Copying data...')
        rind=1
        cind=1
        for th in soup.findAll('th'):
            ws.cell(rind,cind,th.string)
            ws.cell(rind,cind).font=Font(bold=True)
            cind+=1
        for tr in soup.findAll('tr'):
            cind=1
            for td in tr.findChildren('td'):
                a=td.findChild('a')
                if(a!=None):
                    ws.cell(rind,cind,a.string)
                    ws.cell(rind,cind).hyperlink=a['href']
                    ws.cell(rind,cind).style='Hyperlink'
                else:
                    ws.cell(rind,cind,td.string)
                cind+=1
            rind+=1
        ws.delete_cols(1,1)
        wb.save(dest)
        click.echo('Finished copying!')

dumptoexcel()