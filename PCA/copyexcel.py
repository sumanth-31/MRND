from openpyxl import load_workbook,Workbook
from copy import copy
from os.path import isfile
import click


@click.command()
@click.option('-cp/-ncp','--capitalize/--no-capitalize',default=False,help='capitalizes data')
@click.option('-ps/-nps','--preservestyle/--no-preservestyle',default=False,help='Preserves style')
@click.argument('src',nargs=1)
@click.argument('dest',nargs=1)
def copyexcel(capitalize,preservestyle,src,dest):
    '''Command Line Tool For Copying Excel Files'''
    if(isfile(dest)):
        if(not click.confirm('File already exists. Do you want to overwrite?')):
            return
    click.echo('Copying files...')
    wbsrc=load_workbook(src,data_only=preservestyle)
    wslists=wbsrc.sheetnames
    wbdest=Workbook()
    for sheetname in wslists:
        srcsheet=wbsrc[sheetname]
        destsheet=wbdest.create_sheet(sheetname)
        for rowind in range(1,srcsheet.max_row+1):
            destsheet._add_row()
            for colind in range(1,srcsheet.max_column+1):
                cell=srcsheet.cell(rowind,colind)
                destsheet.cell(rowind,colind,cell.value.upper() if capitalize else cell.value)
                if(preservestyle):
                    new_cell=destsheet.cell(rowind,colind)
                    preserveStyle(cell,new_cell)
        if(sheetname=='Current'):
            for i in range(1,destsheet.max_row+1):
                destsheet.cell(i,5,'ol2016_'+destsheet.cell(i,2).value+'_'+destsheet.cell(i,4).value+'_mock')
        for idx,rd in srcsheet.row_dimensions.items():
            destsheet.row_dimensions[idx]=copy(rd)
        for idx,cd in srcsheet.column_dimensions.items():
            destsheet.column_dimensions[idx]=copy(cd)
    wsdestlist=wbdest.sheetnames
    for sheetname in wsdestlist:
        if(sheetname not in wslists):
            wbdest.remove(wbdest[sheetname])
    wbdest.save(dest)
    click.echo('Finished copying!')

def preserveStyle(cell,new_cell):
    if cell.has_style:
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.protection = copy(cell.protection)
        new_cell.alignment = copy(cell.alignment)


copyexcel()
