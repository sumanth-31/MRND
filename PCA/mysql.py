import MySQLdb as sql
import click
from openpyxl import Workbook,load_workbook
import logging


conn=None
cursor=None
logger=None

@click.group()
def main():
    global cursor,conn,logger
    conn=sql.connect(host='localhost',port=8082,user='root',passwd='root')
    conn.autocommit(True)
    cursor=conn.cursor()
    loggerStreamHandler=logging.StreamHandler()
    loggerStreamHandler.setLevel(logging.WARNING)
    logger=logging.getLogger(__name__)
    logger.addHandler(loggerStreamHandler)
    logger.setLevel(logging.WARNING)
    logger.propagate=False





@main.command()
def createdb():
    global cursor,logger,conn
    try:
        cursor.execute("create database mrnd")
        conn=sql.connect(host='localhost',user='root',port=8082,passwd='root',db='mrnd')
        conn.autocommit(True)
        cursor=conn.cursor()
        cursor.execute("create table students(name varchar(60),college varchar(60),emailid varchar(60),dbnames varchar(20),pkrow varchar(60) primary key)")
        cursor.execute("create table marks(student varchar(60),marks int(8), foreign key (student) references students(pkrow))");
    except Exception as e:
        print(e)
        logger.error('ERROR: Failed to create database!')
        print('error')
    

@main.command()
def dropdb():
    global cursor,logger,conn
    try:
        cursor.execute("drop database mrnd;")
    except:
        logger.error('ERROR: Failed to drop database!')

@main.command()
def importdata():
    conn=sql.connect(host='localhost',user='root',port=8082,passwd='root',db='mrnd')
    conn.autocommit(True)
    wbstudents=load_workbook("students.xlsx")
    wsstudents=wbstudents['Current']
    cursor=conn.cursor()
    for row in range(2,wsstudents.max_row+1):
        arr=[]
        for col in range(1,wsstudents.max_column+1):
            arr.append(wsstudents.cell(row,col).value)
        cursor.execute('insert into students values(%s, %s, %s, %s, %s)',arr)
    wbmarks=load_workbook("marks.xlsx")
    wsmarks=wbmarks['Sheet']
    for row in range(2,wsmarks.max_row+1):
        arr=[wsmarks.cell(row,1).value,wsmarks.cell(row,6).value]
        cursor.execute('insert into marks values(%s, %s)',arr)
    

@main.command()
def collegestats():
    conn=sql.connect(user='root',passwd='root',port=8082,db='mrnd',host='localhost')
    conn.autocommit(True)
    cursor=conn.cursor()
    query='select students.college as college , min(m.marks) as min , max(m.marks) as max , avg(marks) as average from students , marks m where students.pkrow=m.student group by students.college;'
    cursor.execute(query)
    arr=cursor.fetchall()
    for name in ['College','Min','Max','Average']:
        print(name,end='\t')
    print()
    for row in arr:
        for col in row:
            print(col,end='\t')
        print()

main()